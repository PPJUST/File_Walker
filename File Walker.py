import base64
import os
import sys

from PySide2.QtGui import QPixmap, QIcon
from PySide2.QtWidgets import QApplication, QFileDialog
from PySide2.QtWidgets import QMainWindow
from openpyxl import Workbook, load_workbook, styles

from icon_base64 import icon_base64
from ui import Ui_Form


class FileWalker(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # 连接槽函数与信号
        self.ui.button_page_main.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(0))  # 切换页面
        self.ui.button_page_setting.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(1))  # 切换页面
        self.ui.button_quit.clicked.connect(lambda: sys.exit(1))
        self.ui.button_walk.clicked.connect(self.start_walk)
        self.ui.button_ask_path.clicked.connect(
            lambda: self.ui.lineedit_path.setText(QFileDialog.getExistingDirectory(self, "选择文件夹")))
        self.ui.button_open_path.clicked.connect(lambda: os.startfile(self.ui.lineedit_path.text()))
        self.ui.button_open_result.clicked.connect(lambda: os.startfile('遍历结果.xlsx'))

    def start_walk(self):
        """开始遍历"""
        folder = self.ui.lineedit_path.text()  # 获取要遍历的路径
        if folder and os.path.exists(folder):
            self.ui.button_open_result.setEnabled(False)  # 关闭按钮
            print('测试 断点1')
            walk_result = self.walk_files_with_folder(folder)
            print('测试 断点2')
            walk_result_with_size = self.get_folder_size(folder, walk_result)
            print('测试 断点3')
            self.save_xlsx(walk_result_with_size)
            print('测试 断点4')
            if self.ui.checkbox_hyperlink.isChecked():  # 检查超链接按钮
                self.add_hyperlink()
            self.check_size()  # 检查显示文件大小按钮
            self.ui.button_open_result.setEnabled(True)  # 开启按钮

    def save_xlsx(self, walk_result):
        """保存结果为xlsx文件"""
        print('测试：执行 save_xlsx')
        wb = Workbook()
        wb.save('遍历结果.xlsx')  # 新建或覆盖xlsx文件

        wb = load_workbook('遍历结果.xlsx')
        ws = wb.active
        all_files = [key for key in walk_result.values() if key['type'] == '文件']
        total_size = 0
        for item in all_files:
            total_size += item['size_byte']
        if self.ui.checkbox_folder.isChecked():
            ws['A1'] = '文件名'
            ws['B1'] = '文件路径'
            ws['C1'] = '文件大小'
            ws['D1'] = '类型（文件/文件夹）'
            ws['A2'] = os.path.split(self.ui.lineedit_path.text())[1]
            ws['B2'] = self.ui.lineedit_path.text()
            ws['C2'] = total_size
            ws['D2'] = '文件夹'
            write_line = 3
            for key in walk_result:
                ws[f'A{write_line}'] = walk_result[key]['filename']
                ws[f'B{write_line}'] = walk_result[key]['filepath']
                ws[f'C{write_line}'] = walk_result[key]['size_byte']
                ws[f'D{write_line}'] = walk_result[key]['type']
                write_line += 1
        else:
            ws['A1'] = '文件名'
            ws['B1'] = '文件路径'
            ws['C1'] = '文件大小'
            ws['A2'] = os.path.split(self.ui.lineedit_path.text())[1]
            ws['B2'] = self.ui.lineedit_path.text()
            ws['C2'] = total_size
            write_line = 3
            for key in walk_result:
                if os.path.isfile(key):
                    ws[f'A{write_line}'] = walk_result[key]['filename']
                    ws[f'B{write_line}'] = walk_result[key]['filepath']
                    ws[f'C{write_line}'] = walk_result[key]['size_byte']
                    write_line += 1
        wb.save('遍历结果.xlsx')

    @staticmethod
    def add_hyperlink():
        """添加超链接"""
        print('测试：执行 add_hyperlink')
        wb = load_workbook('遍历结果.xlsx')
        ws = wb.active
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            ws[f'B{row}'].hyperlink = ws[f'B{row}'].value
            if os.path.isfile(ws[f'B{row}'].value):
                ws[f'B{row}'].font = styles.Font(color='0000FF', underline='single')
            else:
                ws[f'B{row}'].font = styles.Font(color='FFC125', underline='single')
        wb.save('遍历结果.xlsx')

    def check_size(self):
        """检查显示文件大小按钮"""
        wb = load_workbook('遍历结果.xlsx')
        ws = wb.active
        max_row = ws.max_row
        if self.ui.checkbox_size.isChecked():
            byte_type = self.ui.combobox_byte.currentText()
            if byte_type == 'B':
                ws['C1'] = '文件大小（B）'
            elif byte_type == 'KB':
                ws['C1'] = '文件大小（KB）'
                for row in range(2, max_row + 1):
                    ws[f'C{row}'] = round(ws[f'C{row}'].value / 1024, 2)
            elif byte_type == 'MB':
                ws['C1'] = '文件大小（MB）'
                for row in range(2, max_row + 1):
                    ws[f'C{row}'] = round(ws[f'C{row}'].value / 1024 / 1024, 2)
            elif byte_type == 'GB':
                ws['C1'] = '文件大小（GB）'
                for row in range(2, max_row + 1):
                    ws[f'C{row}'] = round(ws[f'C{row}'].value / 1024 / 1024 / 1024, 2)
        else:
            ws.delete_rows(3)  # 删除"文件大小"列
        wb.save('遍历结果.xlsx')

    @staticmethod
    def walk_files_with_folder(folder):
        """遍历所有文件，显示文件夹"""
        """函数逻辑：
        os.listdir遍历文件夹，如果是文件则返回，如果是文件夹则递归
        保存文件结构：文件名、文件路径、文件大小、是文件还是文件夹"""
        print('测试：执行 walk_files_with_folder')
        walk_result = {}

        def loop_walk(the_folder):
            try:
                walk_filename = os.listdir(the_folder)
                walk_filepath = [os.path.join(the_folder, x).replace('\\', '/') for x in walk_filename]
                walk_filepath_folders = [x for x in walk_filepath if os.path.isdir(x)]  # 提取遍历的文件夹列表
                walk_filepath_files = [x for x in walk_filepath if os.path.isfile(x)]  # 提取遍历的文件列表

                if walk_filepath_folders:  # 如果文件夹列表不为空
                    for item in walk_filepath_folders:
                        if os.path.isdir(item):  # 如果是文件夹，则递归
                            add_info = {'filename': os.path.split(item)[1], 'filepath': item, 'size_byte': 0,
                                        'type': '文件夹'}
                            walk_result[item] = add_info
                            loop_walk(item)
                        else:
                            add_info = {'filename': os.path.split(item)[1], 'filepath': item,
                                        'size_byte': os.path.getsize(item), 'type': '文件'}
                            walk_result[item] = add_info

                if walk_filepath_files:  # 如果文件列表不为空
                    for item in walk_filepath_files:
                        add_info = {'filename': os.path.split(item)[1], 'filepath': item,
                                    'size_byte': os.path.getsize(item), 'type': '文件'}
                        walk_result[item] = add_info

                return walk_result
            except PermissionError:  # 如果无权限访问文件夹会报错PermissionError: [WinError 5] 拒绝访问。
                pass

        walk_result = loop_walk(folder)
        return walk_result

    @staticmethod
    def get_folder_size(the_top_folder, walk_dict):
        """根据遍历所得的字典获取其中文件夹的大小"""
        # 提取文件夹大小的逻辑：取得一个文件路径，将该文件的大小在每一级文件夹中加上
        print('测试：执行 get_folder_size')
        file_in_walk_dict = [key for key in walk_dict if os.path.isfile(key)]  # 提取遍历字典中的文件
        file_in_walk_dict = file_in_walk_dict[::-1]  # 反转列表，实现文件夹层级多的在前面
        get_folder_size_dict = {}
        for item in file_in_walk_dict:
            parent_folder = os.path.split(item)[0]  # 提取的父目录
            while True:  # 循环，获取每级文件夹并加上该文件的大小
                if parent_folder not in get_folder_size_dict:
                    get_folder_size_dict[parent_folder] = 0
                get_folder_size_dict[parent_folder] += walk_dict[item]['size_byte']
                if parent_folder == the_top_folder:
                    break
                else:
                    parent_folder = os.path.split(parent_folder)[0]

        folder_in_walk_dict = [key for key in walk_dict if os.path.isdir(key)]  # 提取遍历字典中的文件夹
        for i in folder_in_walk_dict:
            if i in get_folder_size_dict:  # 如果不在其中，说明是空文件夹，不做处理
                walk_dict[i]['size_byte'] = get_folder_size_dict[i]
        return walk_dict


def base64_to_icon():
    """转换base64为图片用于图标"""
    pixmap = QPixmap()
    pixmap.loadFromData(base64.b64decode(icon_base64))
    icon = QIcon(pixmap)
    return icon


app = QApplication([])
app.setStyle('Fusion')
show_ui = FileWalker()
show_ui.setFixedSize(282, 81)
show_ui.setWindowIcon(QIcon(base64_to_icon()))
show_ui.show()
app.exec_()
